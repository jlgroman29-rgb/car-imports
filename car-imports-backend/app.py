from flask import Flask, has_request_context, request, send_from_directory
from datetime import datetime
import base64
from decimal import Decimal, InvalidOperation
import hashlib
import hmac
import json
import os
import psycopg2
import time
import unicodedata
import uuid
import click
from functools import wraps
from flasgger import Swagger
from flask_cors import CORS
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename


def load_env_file(path):
    if not os.path.exists(path):
        return

    with open(path, "r", encoding="utf-8") as env_file:
        for raw_line in env_file:
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue

            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            os.environ.setdefault(key, value)


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_env_file(os.path.join(BASE_DIR, ".env"))

app = Flask(__name__)
app.config["AUTH_TOKEN_SECRET"] = (
    os.environ.get("JWT_SECRET")
    or os.environ.get("AUTH_TOKEN_SECRET")
    or os.environ.get("SECRET_KEY")
    or "dev-auth-secret-change-me"
)
app.config["AUTH_TOKEN_EXPIRES_SECONDS"] = int(
    os.environ.get("AUTH_TOKEN_EXPIRES_SECONDS", "86400")
)
app.config["FLASK_ENV"] = os.environ.get("FLASK_ENV", "development")
app.config["DB_CONFIG"] = {
    "host": os.environ.get("DB_HOST", "127.0.0.1"),
    "port": int(os.environ.get("DB_PORT", "5433")),
    "database": os.environ.get("DB_NAME", "car_imports"),
    "user": os.environ.get("DB_USER", "postgres"),
    "password": os.environ.get("DB_PASSWORD", "postgres"),
}
app.config["VEHICLE_UPLOAD_FOLDER"] = os.path.join(BASE_DIR, "uploads", "vehicles")
app.config["VEHICLE_IMAGE_MAX_BYTES"] = 5 * 1024 * 1024
app.config["VEHICLE_DOCUMENT_UPLOAD_FOLDER"] = os.path.join(BASE_DIR, "uploads", "vehicle_documents")
app.config["VEHICLE_DOCUMENT_MAX_BYTES"] = 10 * 1024 * 1024
CORS(app)

ALLOWED_VEHICLE_IMAGE_EXTENSIONS = {"jpg", "jpeg", "png", "webp"}
ALLOWED_VEHICLE_DOCUMENT_EXTENSIONS = {"pdf", "jpg", "jpeg", "png", "webp"}

VALID_VEHICLE_DOCUMENT_TYPES = [
    "factura_subasta",
    "titulo",
    "liquidacion_aduana",
    "matricula_dgii",
    "seguro",
    "inspeccion",
    "otros"
]

REQUIRED_VEHICLE_DOCUMENT_TYPES = [
    "factura_subasta",
    "titulo",
    "liquidacion_aduana",
    "matricula_dgii"
]

VALID_ESTADOS = [
    "comprado",
    "en_transito",
    "en_aduana",
    "en_reparacion",
    "disponible",
    "vendido",
    "inventario"
]

VALID_TIPOS_COSTO = [
    "flete",
    "aduana",
    "impuestos",
    "reparacion",
    "transporte_local",
    "comision",
    "documentacion","compra",
    "otros",
]

VALID_QUOTE_STATUSES = ["borrador", "emitida", "cancelada", "vencida", "aprobada", "convertida"]

CUSTOMS_VALUES_SOURCE_YEAR = "2025-2026"
CUSTOMS_ESTIMATE_MODALITIES = {
    "dealer": Decimal("0.10"),
    "particular": Decimal("0.20"),
    "dr_cafta": Decimal("0"),
}
CUSTOMS_VALUE_COLUMN_ALIASES = {
    "marca": "marca",
    "brand": "marca",
    "make": "marca",
    "modelo": "modelo",
    "model": "modelo",
    "ano": "anio",
    "anio": "anio",
    "year": "anio",
    "pais": "pais",
    "pais origen": "pais",
    "pais de origen": "pais",
    "origen": "pais",
    "country": "pais",
    "especificacion": "especificacion_producto",
    "especificaciones": "especificacion_producto",
    "especificacion producto": "especificacion_producto",
    "especificacion de producto": "especificacion_producto",
    "especificacion del producto": "especificacion_producto",
    "descripcion": "especificacion_producto",
    "version": "especificacion_producto",
    "trim": "especificacion_producto",
    "valor": "valor_aduanas",
    "valor aduanas": "valor_aduanas",
    "valor de aduanas": "valor_aduanas",
    "valor aduana": "valor_aduanas",
    "valor aduanal": "valor_aduanas",
    "valor fob": "valor_aduanas",
    "fob": "valor_aduanas",
    "precio": "valor_aduanas",
}
CUSTOMS_VALUE_IGNORED_COLUMN_HEADERS = {
    "codigo arancel",
    "codigo de producto",
    "nombre de producto",
}


def get_connection():
    return psycopg2.connect(**app.config["DB_CONFIG"])


def ensure_users_table(conn):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            email TEXT UNIQUE NOT NULL,
            name TEXT,
            role TEXT NOT NULL DEFAULT 'user',
            password_hash TEXT NOT NULL,
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
    """)
    cur.execute("""
        ALTER TABLE users
        ADD COLUMN IF NOT EXISTS role TEXT NOT NULL DEFAULT 'user';
    """)
    cur.execute("""
        ALTER TABLE users
        ADD COLUMN IF NOT EXISTS is_active BOOLEAN NOT NULL DEFAULT TRUE;
    """)
    cur.execute("""
        ALTER TABLE users
        ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP;
    """)
    conn.commit()
    cur.close()


def ensure_vehicle_media_columns(conn):
    cur = conn.cursor()
    cur.execute("""
        ALTER TABLE vehicles
        ADD COLUMN IF NOT EXISTS color TEXT;
    """)
    cur.execute("""
        ALTER TABLE vehicles
        ADD COLUMN IF NOT EXISTS image_url TEXT;
    """)
    conn.commit()
    cur.close()


def ensure_vehicle_upload_folder():
    os.makedirs(app.config["VEHICLE_UPLOAD_FOLDER"], exist_ok=True)


def ensure_vehicle_document_upload_folder():
    os.makedirs(app.config["VEHICLE_DOCUMENT_UPLOAD_FOLDER"], exist_ok=True)


def ensure_vehicle_documents_table(conn):
    ensure_users_table(conn)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS vehicle_documents (
            id SERIAL PRIMARY KEY,
            vehicle_id INTEGER NOT NULL REFERENCES vehicles(id) ON DELETE CASCADE,
            document_type TEXT NOT NULL,
            original_file_name TEXT NOT NULL,
            stored_file_name TEXT NOT NULL,
            file_path TEXT NOT NULL,
            mime_type TEXT,
            file_size INTEGER NOT NULL,
            notes TEXT,
            uploaded_by INTEGER NULL REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
    """)
    conn.commit()
    cur.close()


def ensure_customs_vehicle_values_table(conn):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS customs_vehicle_values (
            id SERIAL PRIMARY KEY,
            marca TEXT NOT NULL,
            modelo TEXT NOT NULL,
            anio INTEGER NOT NULL,
            pais TEXT,
            especificacion_producto TEXT,
            valor_aduanas NUMERIC(14, 2) NOT NULL,
            source_year TEXT NOT NULL,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
    """)
    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_customs_vehicle_values_unique
        ON customs_vehicle_values (
            LOWER(marca),
            LOWER(modelo),
            anio,
            LOWER(COALESCE(pais, '')),
            LOWER(COALESCE(especificacion_producto, '')),
            source_year
        );
    """)
    conn.commit()
    cur.close()


def normalize_customs_header(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("_", " ").replace("-", " ").replace("/", " ")
    return " ".join(text.split())


def map_customs_column(value):
    normalized = normalize_customs_header(value)
    if normalized in CUSTOMS_VALUE_IGNORED_COLUMN_HEADERS:
        return None

    if normalized in CUSTOMS_VALUE_COLUMN_ALIASES:
        return CUSTOMS_VALUE_COLUMN_ALIASES[normalized]

    for pattern, mapped_name in CUSTOMS_VALUE_COLUMN_ALIASES.items():
        if pattern in normalized:
            return mapped_name

    return None


def normalize_customs_text(value):
    if value is None:
        return None

    text = str(value).strip()
    return " ".join(text.split()) or None


def parse_customs_year(value):
    if value is None or value == "":
        return None

    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def parse_customs_decimal(value):
    if value is None or value == "":
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))

    text = str(value).strip()
    if not text:
        return None

    cleaned = "".join(ch for ch in text if ch.isdigit() or ch in ",.-")
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")

    try:
        return Decimal(cleaned).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def serialize_customs_value(row):
    valor = row.get("valor_aduanas")
    return {
        "id": row.get("id"),
        "marca": row.get("marca"),
        "modelo": row.get("modelo"),
        "anio": row.get("anio"),
        "pais": row.get("pais"),
        "especificacion_producto": row.get("especificacion_producto"),
        "valor_aduanas": float(valor) if valor is not None else None,
        "source_year": row.get("source_year"),
        "created_at": row.get("created_at"),
    }


def decimal_to_float(value):
    if value is None:
        return None

    return float(value)


def parse_required_decimal(data, field_name, allow_zero=True):
    if field_name not in data or data.get(field_name) in (None, ""):
        return None, {
            "status": "error",
            "message": f"{field_name} es requerido"
        }

    value = parse_customs_decimal(data.get(field_name))
    if value is None:
        return None, {
            "status": "error",
            "message": f"{field_name} debe ser numÃ©rico"
        }

    if value < 0 or (not allow_zero and value == 0):
        operator = ">= 0" if allow_zero else "> 0"
        return None, {
            "status": "error",
            "message": f"{field_name} debe ser {operator}"
        }

    return value, None


def parse_optional_decimal(data, field_name, default=Decimal("0")):
    if field_name not in data or data.get(field_name) in (None, ""):
        return default, None

    value = parse_customs_decimal(data.get(field_name))
    if value is None:
        return None, {
            "status": "error",
            "message": f"{field_name} debe ser numÃ©rico"
        }

    if value < 0:
        return None, {
            "status": "error",
            "message": f"{field_name} no puede ser negativo"
        }

    return value, None


def fetch_customs_value_by_id(conn, customs_value_id):
    import psycopg2.extras

    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT id, marca, modelo, anio, pais, especificacion_producto,
               valor_aduanas, source_year, created_at
        FROM customs_vehicle_values
        WHERE id = %s
        LIMIT 1;
    """, (customs_value_id,))
    row = cur.fetchone()
    cur.close()
    return row


def find_customs_value_candidates(conn, data):
    import psycopg2.extras

    marca = normalize_customs_text(data.get("marca"))
    modelo = normalize_customs_text(data.get("modelo"))
    anio = parse_customs_year(data.get("anio"))
    especificacion = normalize_customs_text(data.get("especificacion"))

    missing_fields = []
    if not marca:
        missing_fields.append("marca")
    if not modelo:
        missing_fields.append("modelo")
    if data.get("anio") in (None, ""):
        missing_fields.append("anio")
    if missing_fields:
        return None, {
            "status": "error",
            "message": f"Faltan campos para buscar valor Aduanas: {', '.join(missing_fields)}"
        }

    if anio is None:
        return None, {
            "status": "error",
            "message": "anio debe ser numÃ©rico"
        }

    filters = [
        "marca ILIKE %s",
        "modelo ILIKE %s",
        "anio = %s",
    ]
    params = [marca, modelo, anio]

    if especificacion:
        filters.append("especificacion_producto ILIKE %s")
        params.append(f"%{especificacion}%")

    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"""
        SELECT id, marca, modelo, anio, pais, especificacion_producto,
               valor_aduanas, source_year, created_at
        FROM customs_vehicle_values
        WHERE {" AND ".join(filters)}
        ORDER BY marca, modelo, anio DESC, especificacion_producto NULLS LAST
        LIMIT 25;
    """, tuple(params))
    rows = cur.fetchall()
    cur.close()
    return rows, None


def calculate_customs_modality(cif_usd, tasa_cambio, gravamen_rate, marbete_dop, co2_dop, servicios_aduaneros_dop):
    gravamen_usd = (cif_usd * gravamen_rate).quantize(Decimal("0.01"))
    itbis_base_usd = cif_usd + gravamen_usd
    itbis_usd = (itbis_base_usd * Decimal("0.18")).quantize(Decimal("0.01"))
    subtotal_usd = (cif_usd + gravamen_usd + itbis_usd).quantize(Decimal("0.01"))
    subtotal_dop = (subtotal_usd * tasa_cambio).quantize(Decimal("0.01"))
    total_dop = (subtotal_dop + marbete_dop + co2_dop + servicios_aduaneros_dop).quantize(Decimal("0.01"))

    return {
        "gravamen_usd": decimal_to_float(gravamen_usd),
        "itbis_usd": decimal_to_float(itbis_usd),
        "subtotal_usd": decimal_to_float(subtotal_usd),
        "subtotal_dop": decimal_to_float(subtotal_dop),
        "marbete_dop": decimal_to_float(marbete_dop),
        "co2_dop": decimal_to_float(co2_dop),
        "servicios_aduaneros_dop": decimal_to_float(servicios_aduaneros_dop),
        "total_dop": decimal_to_float(total_dop),
    }


def build_customs_estimate(customs_value, inputs):
    fob_usd = Decimal(str(customs_value["valor_aduanas"])).quantize(Decimal("0.01"))
    seguro_usd = (fob_usd * Decimal("0.02")).quantize(Decimal("0.01"))
    cif_usd = (fob_usd + seguro_usd + inputs["flete_usd"]).quantize(Decimal("0.01"))

    modalidades = {}
    for name, gravamen_rate in CUSTOMS_ESTIMATE_MODALITIES.items():
        modalidades[name] = calculate_customs_modality(
            cif_usd,
            inputs["tasa_cambio"],
            gravamen_rate,
            inputs["marbete_dop"],
            inputs["co2_dop"],
            inputs["servicios_aduaneros_dop"],
        )

    return {
        "customs_value": serialize_customs_value(customs_value),
        "inputs": {
            "tasa_cambio": decimal_to_float(inputs["tasa_cambio"]),
            "fob_usd": decimal_to_float(fob_usd),
            "flete_usd": decimal_to_float(inputs["flete_usd"]),
            "seguro_usd": decimal_to_float(seguro_usd),
            "cif_usd": decimal_to_float(cif_usd),
            "marbete_dop": decimal_to_float(inputs["marbete_dop"]),
            "co2_dop": decimal_to_float(inputs["co2_dop"]),
            "servicios_aduaneros_dop": decimal_to_float(inputs["servicios_aduaneros_dop"]),
        },
        "modalidades": modalidades,
    }


def detect_customs_header(row):
    mapped_columns = {}
    for index, value in enumerate(row):
        mapped_name = map_customs_column(value)
        if mapped_name and mapped_name not in mapped_columns:
            mapped_columns[mapped_name] = index

    required = {"marca", "modelo", "anio", "valor_aduanas"}
    if required.issubset(mapped_columns):
        return mapped_columns

    return None


def row_to_customs_record(row, columns, source_year):
    marca = normalize_customs_text(row[columns["marca"]])
    modelo = normalize_customs_text(row[columns["modelo"]])
    anio = parse_customs_year(row[columns["anio"]])
    valor_aduanas = parse_customs_decimal(row[columns["valor_aduanas"]])

    pais = None
    if "pais" in columns:
        pais = normalize_customs_text(row[columns["pais"]])

    especificacion_producto = None
    if "especificacion_producto" in columns:
        especificacion_producto = normalize_customs_text(row[columns["especificacion_producto"]])

    if not marca or not modelo or anio is None or valor_aduanas is None:
        return None

    return {
        "marca": marca,
        "modelo": modelo,
        "anio": anio,
        "pais": pais,
        "especificacion_producto": especificacion_producto,
        "valor_aduanas": valor_aduanas,
        "source_year": source_year,
    }


def insert_customs_record(cur, record):
    cur.execute("""
        INSERT INTO customs_vehicle_values (
            marca,
            modelo,
            anio,
            pais,
            especificacion_producto,
            valor_aduanas,
            source_year
        )
        SELECT %s, %s, %s, %s, %s, %s, %s
        WHERE NOT EXISTS (
            SELECT 1
            FROM customs_vehicle_values
            WHERE LOWER(marca) = LOWER(%s)
              AND LOWER(modelo) = LOWER(%s)
              AND anio = %s
              AND LOWER(COALESCE(pais, '')) = LOWER(COALESCE(%s, ''))
              AND LOWER(COALESCE(especificacion_producto, '')) = LOWER(COALESCE(%s, ''))
              AND source_year = %s
        )
        RETURNING id;
    """, (
        record["marca"],
        record["modelo"],
        record["anio"],
        record["pais"],
        record["especificacion_producto"],
        record["valor_aduanas"],
        record["source_year"],
        record["marca"],
        record["modelo"],
        record["anio"],
        record["pais"],
        record["especificacion_producto"],
        record["source_year"],
    ))
    return cur.fetchone() is not None


def import_customs_vehicle_values(excel_path, source_year=CUSTOMS_VALUES_SOURCE_YEAR, replace_source_year=False):
    try:
        from openpyxl import load_workbook
    except ImportError as import_error:
        raise RuntimeError("Instala openpyxl para importar archivos .xlsx") from import_error

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Archivo no encontrado: {excel_path}")

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook.active

    columns = None
    data_start_row = None
    scanned_rows = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
        row_values = list(row)
        scanned_rows.append(row_values)
        detected_columns = detect_customs_header(row_values)
        if detected_columns:
            columns = detected_columns
            data_start_row = row_number + 1
            break

    if not columns:
        workbook.close()
        raise ValueError("No se encontraron columnas requeridas: marca, modelo, anio y valor_aduanas")

    conn = get_connection()
    ensure_customs_vehicle_values_table(conn)
    cur = conn.cursor()
    deleted_existing = 0
    if replace_source_year:
        cur.execute("DELETE FROM customs_vehicle_values WHERE source_year = %s;", (source_year,))
        deleted_existing = cur.rowcount

    summary = {
        "status": "OK",
        "source_year": source_year,
        "file": os.path.basename(excel_path),
        "sheet": worksheet.title,
        "header_row": data_start_row - 1,
        "replace_source_year": replace_source_year,
        "deleted_existing": deleted_existing,
        "processed": 0,
        "inserted": 0,
        "duplicates": 0,
        "skipped": 0,
    }

    try:
        rows_iter = worksheet.iter_rows(min_row=data_start_row, values_only=True)
        for row in rows_iter:
            summary["processed"] += 1
            row_values = list(row)
            record = row_to_customs_record(row_values, columns, source_year)
            if not record:
                summary["skipped"] += 1
                continue

            if insert_customs_record(cur, record):
                summary["inserted"] += 1
            else:
                summary["duplicates"] += 1

        log_audit(conn, "import", "customs_vehicle_values", source_year, summary)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()
        workbook.close()

    return summary


def get_file_extension(filename):
    if "." not in filename:
        return ""

    return filename.rsplit(".", 1)[1].lower()


def is_allowed_vehicle_image(filename):
    return get_file_extension(filename) in ALLOWED_VEHICLE_IMAGE_EXTENSIONS


def is_allowed_vehicle_document(filename):
    return get_file_extension(filename) in ALLOWED_VEHICLE_DOCUMENT_EXTENSIONS


def build_vehicle_image_url(filename):
    return f"{request.host_url.rstrip('/')}/uploads/vehicles/{filename}"


def serialize_vehicle_document(row):
    return {
        "id": row.get("id"),
        "vehicle_id": row.get("vehicle_id"),
        "document_type": row.get("document_type"),
        "original_file_name": row.get("original_file_name"),
        "stored_file_name": row.get("stored_file_name"),
        "file_path": row.get("file_path"),
        "mime_type": row.get("mime_type"),
        "file_size": row.get("file_size"),
        "notes": row.get("notes"),
        "uploaded_by": row.get("uploaded_by"),
        "created_at": row.get("created_at"),
        "download_url": f"{request.host_url.rstrip('/')}/vehicle-documents/{row.get('id')}/download"
    }


VALID_USER_ROLES = ["admin", "user"]


def normalize_email(email):
    return (email or "").strip().lower()


def serialize_user(row):
    return {
        "id": row["id"],
        "email": row["email"],
        "name": row.get("name"),
        "role": row.get("role", "user"),
        "is_active": row["is_active"],
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    }


def parse_boolean(value, field_name):
    if isinstance(value, bool):
        return value, None

    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in ("true", "1", "yes", "si", "sí"):
            return True, None
        if normalized in ("false", "0", "no"):
            return False, None

    return None, {
        "status": "error",
        "message": f"{field_name} debe ser booleano"
    }


def validate_user_role(role):
    if role not in VALID_USER_ROLES:
        return {
            "status": "error",
            "message": f"role invalido. Usa uno de estos: {VALID_USER_ROLES}"
        }

    return None


def base64url_encode(data):
    return base64.urlsafe_b64encode(data).rstrip(b"=").decode("ascii")


def base64url_decode(value):
    padding = "=" * (-len(value) % 4)
    return base64.urlsafe_b64decode((value + padding).encode("ascii"))


def create_auth_token(user):
    now = int(time.time())
    expires_in = app.config["AUTH_TOKEN_EXPIRES_SECONDS"]
    header = {"alg": "HS256", "typ": "JWT"}
    payload = {
        "sub": str(user["id"]),
        "email": user["email"],
        "name": user.get("name"),
        "role": user.get("role", "user"),
        "iat": now,
        "exp": now + expires_in,
    }

    header_part = base64url_encode(json.dumps(header, separators=(",", ":")).encode("utf-8"))
    payload_part = base64url_encode(json.dumps(payload, separators=(",", ":")).encode("utf-8"))
    signing_input = f"{header_part}.{payload_part}".encode("ascii")
    signature = hmac.new(
        app.config["AUTH_TOKEN_SECRET"].encode("utf-8"),
        signing_input,
        hashlib.sha256,
    ).digest()

    return f"{header_part}.{payload_part}.{base64url_encode(signature)}", expires_in


def decode_auth_token(token):
    try:
        header_part, payload_part, signature_part = token.split(".")
    except ValueError:
        return None, "Token invalido"

    signing_input = f"{header_part}.{payload_part}".encode("ascii")
    expected_signature = hmac.new(
        app.config["AUTH_TOKEN_SECRET"].encode("utf-8"),
        signing_input,
        hashlib.sha256,
    ).digest()

    try:
        provided_signature = base64url_decode(signature_part)
    except Exception:
        return None, "Token invalido"

    if not hmac.compare_digest(expected_signature, provided_signature):
        return None, "Token invalido"

    try:
        payload = json.loads(base64url_decode(payload_part).decode("utf-8"))
    except Exception:
        return None, "Token invalido"

    if int(payload.get("exp", 0)) < int(time.time()):
        return None, "Token expirado"

    return payload, None


def get_bearer_token():
    auth_header = request.headers.get("Authorization", "")
    parts = auth_header.split()

    if len(parts) == 2 and parts[0].lower() == "bearer":
        return parts[1]

    return None


def get_authenticated_user():
    token = get_bearer_token()
    if not token:
        return None, ({
            "status": "error",
            "message": "Authorization Bearer token requerido"
        }, 401)

    payload, token_error = decode_auth_token(token)
    if token_error:
        return None, ({
            "status": "error",
            "message": token_error
        }, 401)

    try:
        import psycopg2.extras

        conn = get_connection()
        ensure_users_table(conn)
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT id, email, name, role, is_active, created_at, updated_at
            FROM users
            WHERE id = %s
            LIMIT 1;
        """, (payload.get("sub"),))
        user = cur.fetchone()

        cur.close()
        conn.close()
    except Exception as e:
        return None, ({"error": str(e)}, 500)

    if not user:
        return None, ({
            "status": "error",
            "message": "Usuario no encontrado"
        }, 401)

    return user, None


def require_admin_user():
    user, error_response = get_authenticated_user()
    if error_response:
        return None, error_response

    if not user.get("is_active"):
        return None, ({
            "status": "error",
            "message": "Usuario inactivo"
        }, 403)

    if user.get("role") != "admin":
        return None, ({
            "status": "error",
            "message": "Se requiere rol admin"
        }, 403)

    return user, None


def require_admin(handler):
    @wraps(handler)
    def wrapped(*args, **kwargs):
        _, error_response = require_admin_user()
        if error_response:
            return error_response
        return handler(*args, **kwargs)

    return wrapped


def require_authenticated_active_user(handler):
    @wraps(handler)
    def wrapped(*args, **kwargs):
        user, error_response = get_authenticated_user()
        if error_response:
            return error_response

        if not user.get("is_active"):
            return {
                "status": "error",
                "message": "Usuario inactivo"
            }, 403

        return handler(*args, **kwargs)

    return wrapped


def ensure_audit_logs_table(conn):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS audit_logs (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NULL REFERENCES users(id) ON DELETE SET NULL,
            action TEXT NOT NULL,
            entity_type TEXT NOT NULL,
            entity_id TEXT NOT NULL,
            details JSONB,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
    """)
    conn.commit()
    cur.close()


def get_optional_authenticated_user():
    if not has_request_context():
        return None

    token = get_bearer_token()
    if not token:
        return None

    payload, token_error = decode_auth_token(token)
    if token_error:
        return None

    try:
        user_id = int(payload.get("sub"))
    except (TypeError, ValueError):
        return None

    return user_id


def log_audit(conn, action, entity_type, entity_id, details=None):
    ensure_users_table(conn)
    ensure_audit_logs_table(conn)
    user_id = get_optional_authenticated_user()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO audit_logs (user_id, action, entity_type, entity_id, details)
        VALUES (%s, %s, %s, %s, %s::jsonb);
        """,
        (user_id, action, entity_type, str(entity_id), json.dumps(details or {}))
    )
    cur.close()



MIN_ALLOWED_RECORD_YEAR = 2000


def validate_transaction_date(value, field_name="fecha"):
    if value in (None, ""):
        return None, None

    try:
        parsed_date = datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None, ({
            "status": "error",
            "message": f"{field_name} debe tener formato YYYY-MM-DD"
        }, 400)

    today = datetime.utcnow().date()
    if parsed_date > today:
        return None, ({
            "status": "error",
            "message": f"{field_name} no puede ser una fecha futura"
        }, 400)

    if parsed_date.year < MIN_ALLOWED_RECORD_YEAR:
        return None, ({
            "status": "error",
            "message": f"{field_name} debe estar entre {MIN_ALLOWED_RECORD_YEAR} y {today.year}"
        }, 400)

    return parsed_date, None

def parse_date_filter(value, field_name):
    if not value:
        return None, None

    try:
        return datetime.strptime(value, "%Y-%m-%d").date(), None
    except ValueError:
        return None, {
            "status": "error",
            "message": f"{field_name} debe tener formato YYYY-MM-DD"
        }


def get_date_filters():
    start_date, start_error = parse_date_filter(request.args.get("start_date"), "start_date")
    if start_error:
        return None, None, (start_error, 400)

    end_date, end_error = parse_date_filter(request.args.get("end_date"), "end_date")
    if end_error:
        return None, None, (end_error, 400)

    if start_date and end_date and start_date > end_date:
        return None, None, ({
            "status": "error",
            "message": "start_date no puede ser posterior a end_date"
        }, 400)

    return start_date, end_date, None


def build_date_filter_clause(column_name, start_date, end_date, table_alias=None):
    if not column_name:
        return "", []

    qualified_column = f"{table_alias}.{column_name}" if table_alias else column_name
    filters = []
    params = []

    if start_date:
        filters.append(f"{qualified_column}::date >= %s")
        params.append(start_date)

    if end_date:
        filters.append(f"{qualified_column}::date <= %s")
        params.append(end_date)

    if not filters:
        return "", []

    return " AND " + " AND ".join(filters), params

def map_vehicle(row):
    color = None
    image_url = None

    if hasattr(row, "get"):
        color = row.get("color")
        image_url = row.get("image_url")
    else:
        color = row[11] if len(row) > 11 else None
        image_url = row[12] if len(row) > 12 else None

    return {
        "id": row[0],
        "vin": row[1],
        "marca": row[2],
        "modelo": row[3],
        "anio": row[4],
        "estado": row[5],
        "fecha_compra": row[6],
        "fecha_llegada": row[7],
        "precio_estimado": float(row[8]) if row[8] else 0,
        "fecha_venta": row[9],
        "created_at": row[10],
        "color": color,
        "image_url": image_url
    }

def validar_estado(data):
    estado = data.get("estado")

    if estado and estado not in VALID_ESTADOS:
        return {
            "status": "error",
            "message": f"Estado inválido. Usa uno de estos: {VALID_ESTADOS}"
        }, 400

    return None




def format_cost_type_label(cost_type):
    return cost_type.replace("_", " ").title()


def serialize_cost_types_dropdown(cost_types):
    return [
        {"value": cost_type, "label": format_cost_type_label(cost_type)}
        for cost_type in cost_types
    ]

def validar_tipo_costo(data):
    tipo = data.get("tipo")

    if tipo and tipo not in VALID_TIPOS_COSTO:
        return {
            "status": "error",
            "message": f"Tipo de costo inválido. Usa uno de estos: {VALID_TIPOS_COSTO}"
        }, 400

    return None


def get_table_columns(conn, table_name):
    cur = conn.cursor()
    cur.execute("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s
    """, (table_name,))
    cols = {row[0] for row in cur.fetchall()}
    cur.close()
    return cols


def get_sales_price_column(columns):
    if "monto" in columns:
        return "monto"
    if "precio_venta" in columns:
        return "precio_venta"
    return None


def get_sales_date_column(columns):
    if "fecha_venta" in columns:
        return "fecha_venta"
    if "fecha" in columns:
        return "fecha"
    return None


def normalize_sale(row, columns):
    price_column = get_sales_price_column(columns)
    date_column = get_sales_date_column(columns)
    amount = row.get(price_column) if price_column else None
    tasa = row.get("tasa_cambio")
    return {
        "id": row.get("id"),
        "vehicle_id": row.get("vehicle_id"),
        "precio_venta": float(amount) if amount is not None else 0,
        "monto": float(amount) if amount is not None else 0,
        "moneda": row.get("moneda", "DOP"),
        "tasa_cambio": float(tasa) if tasa is not None else None,
        "fecha_venta": row.get(date_column) if date_column else None,
        "nombre_cliente": row.get("nombre_cliente"),
        "telefono_cliente": row.get("telefono_cliente"),
        "metodo_pago": row.get("metodo_pago"),
        "notas": row.get("notas"),
        "fecha": row.get(date_column) if date_column else None,
        "created_at": row.get("created_at"),
    }


def serialize_profit_row(row):
    total_costos = float(row["total_costos"]) if row["total_costos"] is not None else 0.0
    total_venta = float(row["total_venta"]) if row["total_venta"] is not None else 0.0
    ganancia_real = float(row["ganancia_real"]) if row["ganancia_real"] is not None else 0.0
    margen = (ganancia_real / total_venta * 100) if total_venta > 0 else 0.0

    return {
        "vehicle_id": row["vehicle_id"],
        "vin": row["vin"],
        "marca": row["marca"],
        "modelo": row["modelo"],
        "anio": row["anio"],
        "estado": row["estado"],
        "precio_estimado": float(row["precio_estimado"]) if row.get("precio_estimado") is not None else 0.0,
        "total_costos": total_costos,
        "total_venta": total_venta,
        "ganancia_real": ganancia_real,
        "margen_porcentaje": margen,
    }


def ensure_quotes_table(conn):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS quotes (
            id SERIAL PRIMARY KEY,
            vehicle_id INTEGER NOT NULL REFERENCES vehicles(id),
            customer_name TEXT,
            customer_document TEXT,
            customer_phone TEXT,
            customer_email TEXT,
            customer_address TEXT,
            finance_entity TEXT,
            price_usd NUMERIC(12, 2) NOT NULL,
            exchange_rate NUMERIC(12, 4) NOT NULL,
            price_dop NUMERIC(14, 2) NOT NULL,
            valid_until DATE,
            notes TEXT,
            status TEXT NOT NULL DEFAULT 'emitida',
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
    """)
    cur.execute("""
        ALTER TABLE quotes
        ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP;
    """)
    conn.commit()
    cur.close()


def parse_number(value, field_name):
    if value in (None, ""):
        return None, {
            "status": "error",
            "message": f"{field_name} es obligatorio"
        }

    try:
        return float(value), None
    except (TypeError, ValueError):
        return None, {
            "status": "error",
            "message": f"{field_name} debe ser numerico"
        }


def validate_quote_date(value, field_name="valid_until"):
    if value in (None, ""):
        return None, None

    try:
        return datetime.strptime(value, "%Y-%m-%d").date(), None
    except ValueError:
        return None, {
            "status": "error",
            "message": f"{field_name} debe tener formato YYYY-MM-DD"
        }


def validate_quote_status(status):
    if status not in VALID_QUOTE_STATUSES:
        return {
            "status": "error",
            "message": f"status invalido. Usa uno de estos: {VALID_QUOTE_STATUSES}"
        }

    return None


def serialize_quote(row):
    return {
        "id": row.get("id"),
        "vehicle_id": row.get("vehicle_id"),
        "customer_name": row.get("customer_name"),
        "customer_document": row.get("customer_document"),
        "customer_phone": row.get("customer_phone"),
        "customer_email": row.get("customer_email"),
        "customer_address": row.get("customer_address"),
        "finance_entity": row.get("finance_entity"),
        "price_usd": float(row["price_usd"]) if row.get("price_usd") is not None else None,
        "exchange_rate": float(row["exchange_rate"]) if row.get("exchange_rate") is not None else None,
        "price_dop": float(row["price_dop"]) if row.get("price_dop") is not None else None,
        "valid_until": row.get("valid_until"),
        "notes": row.get("notes"),
        "status": row.get("status"),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    }


@app.route("/auth/login", methods=["POST"])
def login():
    data = request.get_json(silent=True) or {}
    email = normalize_email(data.get("email"))
    password = data.get("password") or ""

    if not email or not password:
        return {
            "status": "error",
            "message": "email y password son requeridos"
        }, 400

    try:
        import psycopg2.extras

        conn = get_connection()
        ensure_users_table(conn)
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT id, email, name, role, password_hash, is_active, created_at, updated_at
            FROM users
            WHERE LOWER(email) = %s
            LIMIT 1;
        """, (email,))
        user = cur.fetchone()

        cur.close()
        conn.close()

        if not user or not user["is_active"] or not check_password_hash(user["password_hash"], password):
            return {
                "status": "error",
                "message": "Credenciales invalidas"
            }, 401

        token, expires_in = create_auth_token(user)

        return {
            "status": "OK",
            "token_type": "Bearer",
            "access_token": token,
            "expires_in": expires_in,
            "user": serialize_user(user),
        }

    except Exception as e:
        return {"error": str(e)}, 500


@app.route("/auth/me", methods=["GET"])
def auth_me():
    token = get_bearer_token()
    if not token:
        return {
            "status": "error",
            "message": "Authorization Bearer token requerido"
        }, 401

    payload, token_error = decode_auth_token(token)
    if token_error:
        return {
            "status": "error",
            "message": token_error
        }, 401

    try:
        import psycopg2.extras

        conn = get_connection()
        ensure_users_table(conn)
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT id, email, name, role, is_active, created_at, updated_at
            FROM users
            WHERE id = %s
            LIMIT 1;
        """, (payload.get("sub"),))
        user = cur.fetchone()

        cur.close()
        conn.close()

        if not user or not user["is_active"]:
            return {
                "status": "error",
                "message": "Usuario no encontrado o inactivo"
            }, 401

        return {
            "status": "OK",
            "user": serialize_user(user),
        }

    except Exception as e:
        return {"error": str(e)}, 500


@app.route("/users", methods=["GET"])
@require_admin
def list_users():
    try:
        import psycopg2.extras

        conn = get_connection()
        ensure_users_table(conn)
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT id, email, name, role, is_active, created_at, updated_at
            FROM users
            ORDER BY id ASC;
        """)
        users = cur.fetchall()

        cur.close()
        conn.close()

        return {
            "status": "OK",
            "data": [serialize_user(user) for user in users],
        }

    except Exception as e:
        return {"error": str(e)}, 500


@app.route("/audit-logs", methods=["GET"])
@require_admin
def list_audit_logs():
    try:
        import psycopg2.extras

        conn = get_connection()
        ensure_users_table(conn)
        ensure_audit_logs_table(conn)
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT id, user_id, action, entity_type, entity_id, details, created_at
            FROM audit_logs
            ORDER BY created_at DESC, id DESC
            LIMIT 500;
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return {"status": "OK", "data": rows}
    except Exception as e:
        return {"error": str(e)}, 500


@app.route("/users", methods=["POST"])
@require_admin
def create_user():
    data = request.get_json(silent=True) or {}
    email = normalize_email(data.get("email"))
    password = data.get("password") or ""
    name = data.get("name")
    role = data.get("role", "user")
    is_active = data.get("is_active", True)

    if not email or not password:
        return {
            "status": "error",
            "message": "email y password son requeridos"
        }, 400

    role_error = validate_user_role(role)
    if role_error:
        return role_error, 400

    is_active, is_active_error = parse_boolean(is_active, "is_active")
    if is_active_error:
        return is_active_error, 400

    try:
        import psycopg2.extras
        from psycopg2 import errors

        conn = get_connection()
        ensure_users_table(conn)
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            INSERT INTO users (email, name, role, password_hash, is_active)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id, email, name, role, is_active, created_at, updated_at;
        """, (email, name, role, generate_password_hash(password), is_active))
        user = cur.fetchone()
        log_audit(conn, "create", "user", user["id"], {"payload": {"email": email, "name": name, "role": role, "is_active": is_active}})
        conn.commit()

        cur.close()
        conn.close()

        return {
            "status": "success",
            "message": "Usuario creado correctamente",
            "user": serialize_user(user),
        }, 201

    except errors.UniqueViolation:
        conn.rollback()
        cur.close()
        conn.close()
        return {
            "status": "error",
            "message": "Ya existe un usuario con ese email"
        }, 409
    except Exception as e:
        return {"error": str(e)}, 500


@app.route("/users/<int:user_id>", methods=["PATCH"])
@require_admin
def update_user(user_id):
    data = request.get_json(silent=True) or {}
    allowed_fields = {"name", "email", "role", "is_active"}
    provided_fields = allowed_fields.intersection(data.keys())

    if not provided_fields:
        return {
            "status": "error",
            "message": "Envia al menos uno de estos campos: name, email, role, is_active"
        }, 400

    updates = []
    params = []

    if "name" in data:
        updates.append("name = %s")
        params.append(data.get("name"))

    if "email" in data:
        email = normalize_email(data.get("email"))
        if not email:
            return {
                "status": "error",
                "message": "email no puede estar vacio"
            }, 400
        updates.append("email = %s")
        params.append(email)

    if "role" in data:
        role = data.get("role")
        role_error = validate_user_role(role)
        if role_error:
            return role_error, 400
        updates.append("role = %s")
        params.append(role)

    if "is_active" in data:
        is_active, is_active_error = parse_boolean(data.get("is_active"), "is_active")
        if is_active_error:
            return is_active_error, 400
        updates.append("is_active = %s")
        params.append(is_active)

    updates.append("updated_at = CURRENT_TIMESTAMP")
    params.append(user_id)

    try:
        import psycopg2.extras
        from psycopg2 import errors

        conn = get_connection()
        ensure_users_table(conn)
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(f"""
            UPDATE users
            SET {", ".join(updates)}
            WHERE id = %s
            RETURNING id, email, name, role, is_active, created_at, updated_at;
        """, params)
        user = cur.fetchone()

        if not user:
            conn.rollback()
            cur.close()
            conn.close()
            return {
                "status": "error",
                "message": "Usuario no encontrado"
            }, 404

        action = "deactivate" if "is_active" in data and user["is_active"] is False else "update"
        log_audit(conn, action, "user", user_id, {"payload": data})
        conn.commit()
        cur.close()
        conn.close()

        return {
            "status": "OK",
            "message": "Usuario actualizado correctamente",
            "user": serialize_user(user),
        }

    except errors.UniqueViolation:
        conn.rollback()
        cur.close()
        conn.close()
        return {
            "status": "error",
            "message": "Ya existe un usuario con ese email"
        }, 409
    except Exception as e:
        return {"error": str(e)}, 500


@app.route("/users/<int:user_id>/password", methods=["PATCH"])
@require_admin
def update_user_password(user_id):
    data = request.get_json(silent=True) or {}
    password = data.get("password") or ""

    if not password:
        return {
            "status": "error",
            "message": "password es requerido"
        }, 400

    try:
        import psycopg2.extras

        conn = get_connection()
        ensure_users_table(conn)
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            UPDATE users
            SET password_hash = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
            RETURNING id, email, name, role, is_active, created_at, updated_at;
        """, (generate_password_hash(password), user_id))
        user = cur.fetchone()

        if not user:
            conn.rollback()
            cur.close()
            conn.close()
            return {
                "status": "error",
                "message": "Usuario no encontrado"
            }, 404

        conn.commit()
        cur.close()
        conn.close()

        return {
            "status": "OK",
            "message": "Password actualizado correctamente",
            "user": serialize_user(user),
        }

    except Exception as e:
        return {"error": str(e)}, 500


@app.cli.command("create-user")
@click.argument("email")
@click.password_option()
@click.option("--name", default=None, help="Nombre visible del usuario.")
@click.option(
    "--role",
    default="admin",
    type=click.Choice(VALID_USER_ROLES),
    help="Rol del usuario. Usa admin para el primer usuario administrador.",
)
def create_user_command(email, password, name, role):
    conn = get_connection()
    ensure_users_table(conn)
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO users (email, name, role, password_hash)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (email)
        DO UPDATE SET
            name = EXCLUDED.name,
            role = EXCLUDED.role,
            password_hash = EXCLUDED.password_hash,
            is_active = TRUE,
            updated_at = CURRENT_TIMESTAMP
        RETURNING id;
    """, (normalize_email(email), name, role, generate_password_hash(password)))
    user_id = cur.fetchone()[0]
    conn.commit()
    cur.close()
    conn.close()
    click.echo(f"Usuario listo: {normalize_email(email)} (id={user_id}, role={role})")


@app.route("/")
def home():
    return {"message": "API funcionando 🚗"}


@app.route("/test-db")
def test_db():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT version();")
        version = cur.fetchone()
        cur.close()
        conn.close()

        return {"status": "OK", "postgres": version[0]}
    except Exception as e:
        return {"error": str(e)}


@app.cli.command("import-customs-values")
@click.argument("excel_path", type=click.Path(exists=True, dir_okay=False))
@click.option("--source-year", default=CUSTOMS_VALUES_SOURCE_YEAR, show_default=True)
@click.option("--replace-source-year", is_flag=True, help="Elimina primero los registros del source_year indicado.")
def import_customs_values_command(excel_path, source_year, replace_source_year):
    summary = import_customs_vehicle_values(excel_path, source_year, replace_source_year)
    click.echo(json.dumps(summary, ensure_ascii=False, indent=2, default=str))


@app.route("/customs-values", methods=["GET"])
def get_customs_values():
    try:
        conn = get_connection()
        ensure_customs_vehicle_values_table(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        filters = []
        params = []

        marca = normalize_customs_text(request.args.get("marca"))
        modelo = normalize_customs_text(request.args.get("modelo"))
        especificacion = normalize_customs_text(request.args.get("especificacion"))
        anio = request.args.get("anio")

        if marca:
            filters.append("marca ILIKE %s")
            params.append(f"%{marca}%")

        if modelo:
            filters.append("modelo ILIKE %s")
            params.append(f"%{modelo}%")

        if anio:
            parsed_anio = parse_customs_year(anio)
            if parsed_anio is None:
                cur.close()
                conn.close()
                return {"status": "error", "message": "anio debe ser numérico"}, 400

            filters.append("anio = %s")
            params.append(parsed_anio)

        if especificacion:
            filters.append("especificacion_producto ILIKE %s")
            params.append(f"%{especificacion}%")

        limit = request.args.get("limit", "100")
        try:
            limit = min(max(int(limit), 1), 500)
        except ValueError:
            cur.close()
            conn.close()
            return {"status": "error", "message": "limit debe ser numérico"}, 400

        where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""
        cur.execute(f"""
            SELECT id, marca, modelo, anio, pais, especificacion_producto,
                   valor_aduanas, source_year, created_at
            FROM customs_vehicle_values
            {where_clause}
            ORDER BY marca, modelo, anio DESC, especificacion_producto NULLS LAST
            LIMIT %s;
        """, (*params, limit))

        rows = cur.fetchall()
        cur.close()
        conn.close()

        return {
            "status": "OK",
            "count": len(rows),
            "data": [serialize_customs_value(row) for row in rows]
        }

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/customs-estimate", methods=["POST"])
def create_customs_estimate():
    try:
        data = request.get_json(silent=True) or {}

        tasa_cambio, error = parse_required_decimal(data, "tasa_cambio", allow_zero=False)
        if error:
            return error, 400

        flete_usd, error = parse_required_decimal(data, "flete_usd", allow_zero=True)
        if error:
            return error, 400

        marbete_dop, error = parse_optional_decimal(data, "marbete_dop")
        if error:
            return error, 400

        co2_dop, error = parse_optional_decimal(data, "co2_dop")
        if error:
            return error, 400

        servicios_aduaneros_dop, error = parse_optional_decimal(data, "servicios_aduaneros_dop")
        if error:
            return error, 400

        conn = get_connection()
        ensure_customs_vehicle_values_table(conn)

        customs_value = None
        if data.get("customs_value_id") not in (None, ""):
            try:
                customs_value_id = int(data.get("customs_value_id"))
            except (TypeError, ValueError):
                conn.close()
                return {
                    "status": "error",
                    "message": "customs_value_id debe ser numÃ©rico"
                }, 400

            customs_value = fetch_customs_value_by_id(conn, customs_value_id)
            if not customs_value:
                conn.close()
                return {
                    "status": "error",
                    "message": f"customs_value_id {customs_value_id} no existe"
                }, 404
        else:
            candidates, error = find_customs_value_candidates(conn, data)
            if error:
                conn.close()
                return error, 400

            if not candidates:
                conn.close()
                return {
                    "status": "error",
                    "message": "No se encontraron valores de Aduanas para los criterios enviados",
                    "candidates": []
                }, 404

            if len(candidates) > 1:
                conn.close()
                return {
                    "status": "error",
                    "message": "La bÃºsqueda devolviÃ³ mÃºltiples valores. EnvÃ­a customs_value_id o una especificaciÃ³n mÃ¡s precisa.",
                    "candidates": [serialize_customs_value(row) for row in candidates]
                }, 400

            customs_value = candidates[0]

        conn.close()

        estimate = build_customs_estimate(customs_value, {
            "tasa_cambio": tasa_cambio,
            "flete_usd": flete_usd,
            "marbete_dop": marbete_dop,
            "co2_dop": co2_dop,
            "servicios_aduaneros_dop": servicios_aduaneros_dop,
        })

        return {
            "status": "success",
            "data": estimate
        }

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/customs-values/options", methods=["GET"])
def get_customs_values_options():
    try:
        conn = get_connection()
        ensure_customs_vehicle_values_table(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        marca = normalize_customs_text(request.args.get("marca"))
        modelo = normalize_customs_text(request.args.get("modelo"))
        anio_value = request.args.get("anio")
        parsed_anio = None

        if anio_value:
            parsed_anio = parse_customs_year(anio_value)
            if parsed_anio is None:
                cur.close()
                conn.close()
                return {"status": "error", "message": "anio debe ser numérico"}, 400

        cur.execute("""
            SELECT MIN(marca) AS marca
            FROM customs_vehicle_values
            WHERE marca IS NOT NULL
            GROUP BY LOWER(marca)
            ORDER BY marca;
        """)
        marcas = [row["marca"] for row in cur.fetchall()]

        modelos = []
        anios = []
        especificaciones = []

        if marca:
            cur.execute("""
                SELECT MIN(modelo) AS modelo
                FROM customs_vehicle_values
                WHERE marca ILIKE %s
                  AND modelo IS NOT NULL
                GROUP BY LOWER(modelo)
                ORDER BY modelo;
            """, (marca,))
            modelos = [row["modelo"] for row in cur.fetchall()]

        if marca and modelo:
            cur.execute("""
                SELECT DISTINCT anio
                FROM customs_vehicle_values
                WHERE marca ILIKE %s
                  AND modelo ILIKE %s
                ORDER BY anio DESC;
            """, (marca, modelo))
            anios = [row["anio"] for row in cur.fetchall()]

        if marca and modelo and parsed_anio is not None:
            cur.execute("""
                SELECT DISTINCT especificacion_producto
                FROM customs_vehicle_values
                WHERE marca ILIKE %s
                  AND modelo ILIKE %s
                  AND anio = %s
                  AND especificacion_producto IS NOT NULL
                ORDER BY especificacion_producto;
            """, (marca, modelo, parsed_anio))
            especificaciones = [row["especificacion_producto"] for row in cur.fetchall()]

        cur.close()
        conn.close()

        return {
            "status": "OK",
            "data": {
                "marcas": marcas,
                "modelos": modelos,
                "anios": anios,
                "especificaciones": especificaciones
            }
        }

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/vehicles", methods=["GET"])
def get_vehicles():
    try:
        conn = get_connection()
        ensure_vehicle_media_columns(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        cur.execute("""
            SELECT 
                id,
                vin,
                marca,
                modelo,
                anio,
                estado,
                precio_estimado,
                fecha_compra,
                fecha_llegada,
                fecha_venta,
                created_at,
                color,
                image_url
            FROM vehicles
            WHERE 1=1
        """)

        result = cur.fetchall()

        cur.close()
        conn.close()

        return {
            "status": "OK",
            "data": result
        }

    except Exception as e:
        return {"error": str(e)}
# ✅ ESTE VA FUERA (MISMO NIVEL)


@app.route("/vehicles", methods=["POST"])
@require_authenticated_active_user
def create_vehicle():
    try:
        data = request.json

        # 🔥 VALIDACIÓN ESTADO
        validation_error = validar_estado(data)
        if validation_error:
            return validation_error

        conn = get_connection()
        ensure_vehicle_media_columns(conn)
        cur = conn.cursor()

        query = """
        INSERT INTO vehicles (
            vin,
            marca,
            modelo,
            anio,
            estado,
            fecha_compra,
            fecha_llegada,
            precio_estimado,
            fecha_venta,
            color,
            image_url,
            created_at
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, NOW())
        RETURNING id;
        """

        cur.execute(query, (
            data["vin"],
            data["marca"],
            data["modelo"],
            data["anio"],
            data["estado"],
            data.get("fecha_compra"),
            data.get("fecha_llegada"),
            data.get("precio_estimado", 0),
            data.get("fecha_venta"),
            data.get("color"),
            data.get("image_url")
        ))

        vehicle_id = cur.fetchone()[0]
        log_audit(conn, "create", "vehicle", vehicle_id, {"payload": data})

        conn.commit()
        cur.close()
        conn.close()

        return {
            "status": "success",
            "message": "Vehículo creado correctamente",
            "id": vehicle_id
        }, 201

    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }, 500


@app.route("/vehicles/<int:id>", methods=["PUT"])
@require_authenticated_active_user
def update_vehicle(id):
    try:
        data = request.json

        conn = get_connection()
        ensure_vehicle_media_columns(conn)
        cur = conn.cursor()

        query = """
        UPDATE vehicles
        SET marca=%s, modelo=%s, anio=%s, estado=%s, color=%s, image_url=%s
        WHERE id=%s;
        """

        cur.execute(
            query, (data["marca"], data["modelo"],
                    data["anio"], data["estado"], data.get("color"),
                    data.get("image_url"), id)
        )
        if cur.rowcount == 0:
            conn.commit()
            cur.close()
            conn.close()
            return {"error": "Vehículo no encontrado"}, 404
        log_audit(conn, "update", "vehicle", id, {"payload": data})

        conn.commit()
        cur.close()
        conn.close()

        return {"status": "OK", "message": "Vehículo actualizado"}

    except Exception as e:
        return {"error": str(e)}


@app.route("/vehicles/<int:id>", methods=["GET"])
def get_vehicle_by_id(id):
    try:
        conn = get_connection()
        ensure_vehicle_media_columns(conn)
        cur = conn.cursor()

        cur.execute("""
            SELECT
                id,
                vin,
                marca,
                modelo,
                anio,
                estado,
                fecha_compra,
                fecha_llegada,
                precio_estimado,
                fecha_venta,
                created_at,
                color,
                image_url
            FROM vehicles
            WHERE id = %s;
        """, (id,))
        row = cur.fetchone()

        cur.close()
        conn.close()

        if row is None:
            return {"error": "Vehículo no encontrado"}, 404

        return {
            "status": "OK",
            "data": map_vehicle(row)
        }

    except Exception as e:
        return {"error": str(e)}


@app.route("/vehicles/<int:vehicle_id>/image", methods=["POST"])
@require_authenticated_active_user
def upload_vehicle_image(vehicle_id):
    try:
        if request.content_length and request.content_length > app.config["VEHICLE_IMAGE_MAX_BYTES"]:
            return {
                "status": "error",
                "message": "La imagen excede el tamaño máximo permitido de 5 MB"
            }, 413

        if "file" not in request.files:
            return {
                "status": "error",
                "message": "Debes enviar una imagen en el campo file"
            }, 400

        file = request.files["file"]
        if not file or file.filename == "":
            return {
                "status": "error",
                "message": "Debes seleccionar una imagen"
            }, 400

        if not is_allowed_vehicle_image(file.filename):
            return {
                "status": "error",
                "message": "Tipo de imagen no permitido. Usa jpg, jpeg, png o webp"
            }, 400

        conn = get_connection()
        ensure_vehicle_media_columns(conn)
        cur = conn.cursor()
        cur.execute("SELECT id FROM vehicles WHERE id = %s;", (vehicle_id,))
        vehicle = cur.fetchone()

        if not vehicle:
            cur.close()
            conn.close()
            return {"status": "error", "message": "Vehículo no encontrado"}, 404

        ensure_vehicle_upload_folder()
        extension = get_file_extension(secure_filename(file.filename))
        filename = f"vehicle_{vehicle_id}_{int(time.time())}_{uuid.uuid4().hex}.{extension}"
        file_path = os.path.join(app.config["VEHICLE_UPLOAD_FOLDER"], filename)
        file.save(file_path)

        image_url = build_vehicle_image_url(filename)
        cur.execute("""
            UPDATE vehicles
            SET image_url = %s
            WHERE id = %s;
        """, (image_url, vehicle_id))
        log_audit(conn, "update", "vehicle", vehicle_id, {
            "image_url": image_url,
            "filename": filename
        })

        conn.commit()
        cur.close()
        conn.close()

        return {
            "status": "OK",
            "message": "Imagen del vehículo subida correctamente",
            "image_url": image_url,
            "filename": filename
        }

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/uploads/vehicles/<path:filename>", methods=["GET"])
def serve_vehicle_image(filename):
    ensure_vehicle_upload_folder()
    return send_from_directory(app.config["VEHICLE_UPLOAD_FOLDER"], filename)


@app.route("/vehicles/document-summary", methods=["GET"])
@require_authenticated_active_user
def get_vehicle_document_summary():
    try:
        conn = get_connection()
        ensure_vehicle_documents_table(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT
                vehicle_id,
                ARRAY_AGG(DISTINCT document_type) AS document_types
            FROM vehicle_documents
            GROUP BY vehicle_id;
        """)
        rows = cur.fetchall()

        cur.close()
        conn.close()

        return {
            "status": "OK",
            "required_document_types": REQUIRED_VEHICLE_DOCUMENT_TYPES,
            "data": [
                {
                    "vehicle_id": row["vehicle_id"],
                    "document_types": row.get("document_types") or []
                }
                for row in rows
            ]
        }

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/vehicles/<int:vehicle_id>/documents", methods=["POST"])
@require_authenticated_active_user
def upload_vehicle_document(vehicle_id):
    try:
        if request.content_length and request.content_length > app.config["VEHICLE_DOCUMENT_MAX_BYTES"]:
            return {
                "status": "error",
                "message": "El documento excede el tamaño máximo permitido de 10 MB"
            }, 413

        document_type = (request.form.get("document_type") or "").strip()
        notes = request.form.get("notes")

        if document_type not in VALID_VEHICLE_DOCUMENT_TYPES:
            return {
                "status": "error",
                "message": f"document_type inválido. Usa uno de estos: {VALID_VEHICLE_DOCUMENT_TYPES}"
            }, 400

        if "file" not in request.files:
            return {
                "status": "error",
                "message": "Debes enviar un archivo en el campo file"
            }, 400

        file = request.files["file"]
        if not file or file.filename == "":
            return {
                "status": "error",
                "message": "Debes seleccionar un archivo"
            }, 400

        if not is_allowed_vehicle_document(file.filename):
            return {
                "status": "error",
                "message": "Tipo de documento no permitido. Usa pdf, jpg, jpeg, png o webp"
            }, 400

        conn = get_connection()
        ensure_vehicle_documents_table(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT id FROM vehicles WHERE id = %s;", (vehicle_id,))
        vehicle = cur.fetchone()

        if not vehicle:
            cur.close()
            conn.close()
            return {"status": "error", "message": "Vehículo no encontrado"}, 404

        ensure_vehicle_document_upload_folder()
        original_file_name = secure_filename(file.filename)
        extension = get_file_extension(original_file_name)
        stored_file_name = f"vehicle_{vehicle_id}_{document_type}_{int(time.time())}_{uuid.uuid4().hex}.{extension}"
        file_path = os.path.join(app.config["VEHICLE_DOCUMENT_UPLOAD_FOLDER"], stored_file_name)
        file.save(file_path)
        file_size = os.path.getsize(file_path)
        uploaded_by = get_optional_authenticated_user()

        cur.execute("""
            INSERT INTO vehicle_documents (
                vehicle_id,
                document_type,
                original_file_name,
                stored_file_name,
                file_path,
                mime_type,
                file_size,
                notes,
                uploaded_by
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id, vehicle_id, document_type, original_file_name, stored_file_name,
                      file_path, mime_type, file_size, notes, uploaded_by, created_at;
        """, (
            vehicle_id,
            document_type,
            original_file_name,
            stored_file_name,
            file_path,
            file.mimetype,
            file_size,
            notes,
            uploaded_by
        ))
        document = cur.fetchone()
        log_audit(conn, "create", "vehicle_document", document["id"], {
            "vehicle_id": vehicle_id,
            "document_type": document_type,
            "original_file_name": original_file_name,
            "stored_file_name": stored_file_name
        })

        conn.commit()
        cur.close()
        conn.close()

        return {
            "status": "OK",
            "message": "Documento del vehículo subido correctamente",
            "data": serialize_vehicle_document(document)
        }, 201

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/vehicles/<int:vehicle_id>/documents", methods=["GET"])
def get_vehicle_documents(vehicle_id):
    try:
        conn = get_connection()
        ensure_vehicle_documents_table(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT id FROM vehicles WHERE id = %s;", (vehicle_id,))
        vehicle = cur.fetchone()

        if not vehicle:
            cur.close()
            conn.close()
            return {"status": "error", "message": "Vehículo no encontrado"}, 404

        cur.execute("""
            SELECT id, vehicle_id, document_type, original_file_name, stored_file_name,
                   file_path, mime_type, file_size, notes, uploaded_by, created_at
            FROM vehicle_documents
            WHERE vehicle_id = %s
            ORDER BY created_at DESC, id DESC;
        """, (vehicle_id,))
        documents = cur.fetchall()

        cur.close()
        conn.close()

        return {
            "status": "OK",
            "data": [serialize_vehicle_document(document) for document in documents]
        }

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/vehicle-documents/<int:document_id>/download", methods=["GET"])
def download_vehicle_document(document_id):
    try:
        conn = get_connection()
        ensure_vehicle_documents_table(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT id, original_file_name, stored_file_name, file_path
            FROM vehicle_documents
            WHERE id = %s;
        """, (document_id,))
        document = cur.fetchone()
        cur.close()
        conn.close()

        if not document:
            return {"status": "error", "message": "Documento no encontrado"}, 404

        if not os.path.exists(document["file_path"]):
            return {"status": "error", "message": "Archivo físico no encontrado"}, 404

        return send_from_directory(
            app.config["VEHICLE_DOCUMENT_UPLOAD_FOLDER"],
            document["stored_file_name"],
            as_attachment=True,
            download_name=document["original_file_name"]
        )

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/vehicle-documents/<int:document_id>", methods=["DELETE"])
@require_authenticated_active_user
def delete_vehicle_document(document_id):
    try:
        conn = get_connection()
        ensure_vehicle_documents_table(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            DELETE FROM vehicle_documents
            WHERE id = %s
            RETURNING id, vehicle_id, document_type, original_file_name, stored_file_name,
                      file_path, mime_type, file_size, notes, uploaded_by, created_at;
        """, (document_id,))
        document = cur.fetchone()

        if not document:
            conn.commit()
            cur.close()
            conn.close()
            return {"status": "error", "message": "Documento no encontrado"}, 404

        file_deleted = False
        file_delete_error = None
        if document["file_path"] and os.path.exists(document["file_path"]):
            try:
                os.remove(document["file_path"])
                file_deleted = True
            except OSError as delete_error:
                file_delete_error = str(delete_error)

        log_audit(conn, "delete", "vehicle_document", document_id, {
            "vehicle_id": document["vehicle_id"],
            "document_type": document["document_type"],
            "original_file_name": document["original_file_name"],
            "stored_file_name": document["stored_file_name"],
            "file_deleted": file_deleted,
            "file_delete_error": file_delete_error
        })

        conn.commit()
        cur.close()
        conn.close()

        return {
            "status": "OK",
            "message": "Documento eliminado correctamente",
            "file_deleted": file_deleted,
            "file_delete_error": file_delete_error,
            "data": serialize_vehicle_document(document)
        }

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/vehicles/<int:id>", methods=["DELETE"])
@require_authenticated_active_user
def delete_vehicle(id):
    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("DELETE FROM vehicles WHERE id = %s RETURNING id;", (id,))
        deleted = cur.fetchone()
        if deleted is not None:
            log_audit(conn, "delete", "vehicle", id, {"deleted_id": id})

        conn.commit()
        cur.close()
        conn.close()

        if deleted is None:
            return {"error": "Vehículo no encontrado"}, 404

        return {
            "status": "OK",
            "message": f"Vehículo {id} eliminado"
        }

    except Exception as e:
        return {"error": str(e)}


@app.route("/vehicles/<int:id>", methods=["PATCH"])
@require_authenticated_active_user
def patch_vehicle(id):
    try:
        data = request.json

        conn = get_connection()
        ensure_vehicle_media_columns(conn)
        cur = conn.cursor()

        fields = []
        values = []

        # 🔥 Construcción controlada
        if "vin" in data:
            fields.append("vin = %s")
            values.append(data["vin"])

        if "marca" in data:
            fields.append("marca = %s")
            values.append(data["marca"])

        if "modelo" in data:
            fields.append("modelo = %s")
            values.append(data["modelo"])

        if "anio" in data:
            fields.append("anio = %s")
            values.append(data["anio"])

        if "estado" in data:
            fields.append("estado = %s")
            values.append(data["estado"])

        if "color" in data:
            fields.append("color = %s")
            values.append(data["color"])

        if "image_url" in data:
            fields.append("image_url = %s")
            values.append(data["image_url"])

        if "precio_estimado" in data:
            fields.append("precio_estimado = %s")
            values.append(float(data["precio_estimado"]))

        if "fecha_compra" in data:
            fields.append("fecha_compra = %s")
            values.append(data["fecha_compra"])

        if "fecha_llegada" in data:
            fields.append("fecha_llegada = %s")
            values.append(data["fecha_llegada"])

        # ⚠️ Validación
        if not fields:
            return {"error": "No se enviaron datos para actualizar"}, 400

        values.append(id)

        query = f"""
        UPDATE vehicles
        SET {", ".join(fields)}
        WHERE id = %s
        RETURNING id;
        """

        cur.execute(query, tuple(values))
        updated = cur.fetchone()
        if updated is not None:
            log_audit(conn, "update", "vehicle", id, {"payload": data})

        conn.commit()
        cur.close()
        conn.close()

        if updated is None:
            return {"error": "Vehículo no encontrado"}, 404

        return {
            "status": "OK",
            "message": f"Vehículo {id} actualizado correctamente"
        }

    except Exception as e:
        return {"error": str(e)}

@app.route("/dashboard/summary", methods=["GET"])
def dashboard_summary():
    try:
        conn = get_connection()
        cur = conn.cursor()

        query = """
        SELECT
            COUNT(*) as total_vehiculos,
            COUNT(*) FILTER (WHERE estado = 'vendido') as vendidos,
            COUNT(*) FILTER (WHERE estado != 'vendido') as inventario,
            COALESCE(SUM(precio_estimado), 0) as valor_total
        FROM vehicles;
        """

        cur.execute(query)
        result = cur.fetchone()

        cur.close()
        conn.close()

        return {
            "status": "OK",
            "data": {
                "total_vehiculos": result[0],
                "vendidos": result[1],
                "inventario": result[2],
                "valor_total": float(result[3])
            }
        }

    except Exception as e:
        return {"error": str(e)}    


@app.route("/catalogs/cost-types", methods=["GET"])
def get_cost_types():
    response_format = request.args.get("format", "list")

    if response_format == "dropdown":
        return {
            "status": "OK",
            "data": serialize_cost_types_dropdown(VALID_TIPOS_COSTO)
        }

    return {"status": "OK", "data": VALID_TIPOS_COSTO}


@app.route("/quotes", methods=["POST"])
@require_authenticated_active_user
def create_quote():
    try:
        data = request.get_json(silent=True) or {}

        vehicle_id = data.get("vehicle_id")
        if not vehicle_id:
            return {"status": "error", "message": "vehicle_id es obligatorio"}, 400

        price_usd, price_usd_error = parse_number(data.get("price_usd"), "price_usd")
        if price_usd_error:
            return price_usd_error, 400

        exchange_rate, exchange_rate_error = parse_number(data.get("exchange_rate"), "exchange_rate")
        if exchange_rate_error:
            return exchange_rate_error, 400

        if "price_dop" in data and data.get("price_dop") not in (None, ""):
            price_dop, price_dop_error = parse_number(data.get("price_dop"), "price_dop")
            if price_dop_error:
                return price_dop_error, 400
        else:
            price_dop = price_usd * exchange_rate

        valid_until, date_error = validate_quote_date(data.get("valid_until"))
        if date_error:
            return date_error, 400

        status = data.get("status", "emitida")
        status_error = validate_quote_status(status)
        if status_error:
            return status_error, 400

        conn = get_connection()
        ensure_quotes_table(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT id FROM vehicles WHERE id = %s;", (vehicle_id,))
        vehicle = cur.fetchone()
        if not vehicle:
            cur.close()
            conn.close()
            return {"status": "error", "message": f"Vehiculo {vehicle_id} no existe"}, 404

        cur.execute("""
            INSERT INTO quotes (
                vehicle_id,
                customer_name,
                customer_document,
                customer_phone,
                customer_email,
                customer_address,
                finance_entity,
                price_usd,
                exchange_rate,
                price_dop,
                valid_until,
                notes,
                status,
                created_at,
                updated_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
            RETURNING *;
        """, (
            vehicle_id,
            data.get("customer_name"),
            data.get("customer_document"),
            data.get("customer_phone"),
            data.get("customer_email"),
            data.get("customer_address"),
            data.get("finance_entity"),
            price_usd,
            exchange_rate,
            price_dop,
            valid_until,
            data.get("notes"),
            status,
        ))

        quote = cur.fetchone()
        log_audit(conn, "create", "quote", quote["id"], {"payload": data})

        conn.commit()
        cur.close()
        conn.close()

        return {
            "status": "success",
            "message": "Cotizacion creada correctamente",
            "data": serialize_quote(quote)
        }, 201

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/quotes", methods=["GET"])
def get_quotes():
    try:
        conn = get_connection()
        ensure_quotes_table(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT *
            FROM quotes
            ORDER BY created_at DESC, id DESC;
        """)
        rows = cur.fetchall()

        cur.close()
        conn.close()

        return {"status": "success", "data": [serialize_quote(row) for row in rows]}

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/quotes/<int:id>", methods=["GET"])
def get_quote_by_id(id):
    try:
        conn = get_connection()
        ensure_quotes_table(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM quotes WHERE id = %s;", (id,))
        quote = cur.fetchone()

        cur.close()
        conn.close()

        if not quote:
            return {"status": "error", "message": "Cotizacion no encontrada"}, 404

        return {"status": "success", "data": serialize_quote(quote)}

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/vehicles/<int:vehicle_id>/quotes", methods=["GET"])
def get_quotes_by_vehicle(vehicle_id):
    try:
        conn = get_connection()
        ensure_quotes_table(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT id FROM vehicles WHERE id = %s;", (vehicle_id,))
        vehicle = cur.fetchone()
        if not vehicle:
            cur.close()
            conn.close()
            return {"status": "error", "message": f"Vehiculo {vehicle_id} no existe"}, 404

        cur.execute("""
            SELECT *
            FROM quotes
            WHERE vehicle_id = %s
            ORDER BY created_at DESC, id DESC;
        """, (vehicle_id,))
        rows = cur.fetchall()

        cur.close()
        conn.close()

        return {"status": "success", "data": [serialize_quote(row) for row in rows]}

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/quotes/<int:id>", methods=["PATCH"])
@require_authenticated_active_user
def patch_quote(id):
    try:
        data = request.get_json(silent=True) or {}

        if not data:
            return {"status": "error", "message": "No hay datos para actualizar"}, 400

        conn = get_connection()
        ensure_quotes_table(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM quotes WHERE id = %s;", (id,))
        current = cur.fetchone()
        if not current:
            cur.close()
            conn.close()
            return {"status": "error", "message": "Cotizacion no encontrada"}, 404

        fields = []
        values = []
        allowed_text_fields = [
            "customer_name",
            "customer_document",
            "customer_phone",
            "customer_email",
            "customer_address",
            "finance_entity",
            "notes",
        ]

        for field in allowed_text_fields:
            if field in data:
                fields.append(f"{field} = %s")
                values.append(data[field])

        if "vehicle_id" in data:
            cur.execute("SELECT id FROM vehicles WHERE id = %s;", (data["vehicle_id"],))
            vehicle = cur.fetchone()
            if not vehicle:
                cur.close()
                conn.close()
                return {"status": "error", "message": f"Vehiculo {data['vehicle_id']} no existe"}, 404
            fields.append("vehicle_id = %s")
            values.append(data["vehicle_id"])

        if "valid_until" in data:
            valid_until, date_error = validate_quote_date(data.get("valid_until"))
            if date_error:
                cur.close()
                conn.close()
                return date_error, 400
            fields.append("valid_until = %s")
            values.append(valid_until)

        if "status" in data:
            status_error = validate_quote_status(data["status"])
            if status_error:
                cur.close()
                conn.close()
                return status_error, 400
            fields.append("status = %s")
            values.append(data["status"])

        price_usd = float(current["price_usd"])
        exchange_rate = float(current["exchange_rate"])
        price_changed = False

        if "price_usd" in data:
            price_usd, price_usd_error = parse_number(data.get("price_usd"), "price_usd")
            if price_usd_error:
                cur.close()
                conn.close()
                return price_usd_error, 400
            fields.append("price_usd = %s")
            values.append(price_usd)
            price_changed = True

        if "exchange_rate" in data:
            exchange_rate, exchange_rate_error = parse_number(data.get("exchange_rate"), "exchange_rate")
            if exchange_rate_error:
                cur.close()
                conn.close()
                return exchange_rate_error, 400
            fields.append("exchange_rate = %s")
            values.append(exchange_rate)
            price_changed = True

        if "price_dop" in data:
            price_dop, price_dop_error = parse_number(data.get("price_dop"), "price_dop")
            if price_dop_error:
                cur.close()
                conn.close()
                return price_dop_error, 400
            fields.append("price_dop = %s")
            values.append(price_dop)
        elif price_changed:
            fields.append("price_dop = %s")
            values.append(price_usd * exchange_rate)

        if not fields:
            cur.close()
            conn.close()
            return {"status": "error", "message": "No hay datos para actualizar"}, 400

        fields.append("updated_at = NOW()")
        values.append(id)

        cur.execute(f"""
            UPDATE quotes
            SET {", ".join(fields)}
            WHERE id = %s
            RETURNING *;
        """, tuple(values))

        quote = cur.fetchone()
        log_audit(conn, "update", "quote", id, {"payload": data})

        conn.commit()
        cur.close()
        conn.close()

        return {
            "status": "success",
            "message": "Cotizacion actualizada correctamente",
            "data": serialize_quote(quote)
        }

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/quotes/<int:id>", methods=["DELETE"])
@require_authenticated_active_user
def delete_quote(id):
    try:
        conn = get_connection()
        ensure_quotes_table(conn)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            UPDATE quotes
            SET status = 'cancelada',
                updated_at = NOW()
            WHERE id = %s
            RETURNING *;
        """, (id,))
        quote = cur.fetchone()

        if not quote:
            conn.commit()
            cur.close()
            conn.close()
            return {"status": "error", "message": "Cotizacion no encontrada"}, 404

        log_audit(conn, "delete", "quote", id, {"status": "cancelada"})

        conn.commit()
        cur.close()
        conn.close()

        return {
            "status": "success",
            "message": f"Cotizacion {id} cancelada",
            "data": serialize_quote(quote)
        }

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/sales", methods=["POST"])
@require_authenticated_active_user
def create_sale():
    """
    Crea una venta para un vehículo específico.

    Body esperado (JSON):
        - vehicle_id (int, requerido): ID del vehículo a vender.
        - precio_venta (number, requerido si no se envía monto): monto de venta.
        - monto (number, requerido si no se envía precio_venta): alias de precio_venta.
        - moneda (str, opcional): por defecto "DOP".
        - tasa_cambio (number, opcional): tasa aplicada al momento de la venta.
        - fecha_venta (date/datetime string, opcional): fecha de la venta.
        - fecha (date/datetime string, opcional): alias de fecha_venta.
        - nombre_cliente (str, opcional): nombre del cliente.
        - telefono_cliente (str, opcional): teléfono del cliente.
        - metodo_pago (str, opcional): método de pago registrado.
        - notas (str, opcional): observaciones adicionales.

    Validaciones:
        - La tabla `sales` debe contener una columna de monto válida (`monto` o `precio_venta`).
        - `vehicle_id` es obligatorio.
        - Debe enviarse `precio_venta` o `monto`.
        - El vehículo debe existir en `vehicles`.
        - Regla de negocio: un vehículo solo puede tener una venta registrada.

    Respuestas HTTP:
        - 201: venta creada correctamente.
        - 400: faltan campos obligatorios.
        - 404: el vehículo no existe.
        - 409: ya existe una venta para ese vehículo.
        - 500: error interno o inconsistencia de esquema.
    """
    try:
        data = request.get_json(silent=True) or {}

        conn = get_connection()
        sales_columns = get_table_columns(conn, "sales")
        price_column = get_sales_price_column(sales_columns)
        date_column = get_sales_date_column(sales_columns)

        if not price_column:
            conn.close()
            return {"status": "error", "message": "La tabla sales no tiene una columna de monto válida"}, 500

        vehicle_id = data.get("vehicle_id")
        amount = data.get("precio_venta", data.get("monto"))
        moneda = data.get("moneda", "DOP")
        tasa_cambio = data.get("tasa_cambio")
        fecha_venta = data.get("fecha_venta", data.get("fecha"))
        _, date_error = validate_transaction_date(fecha_venta, "fecha_venta")
        if date_error:
            conn.close()
            return date_error

        nombre_cliente = data.get("nombre_cliente")
        telefono_cliente = data.get("telefono_cliente")
        metodo_pago = data.get("metodo_pago")
        notas = data.get("notas")

        if not vehicle_id:
            conn.close()
            return {"status": "error", "message": "vehicle_id es obligatorio"}, 400

        if amount is None:
            conn.close()
            return {"status": "error", "message": "precio_venta es obligatorio"}, 400

        cur = conn.cursor()

        # Validación referencial: la venta solo puede asociarse a un vehículo existente.
        cur.execute("SELECT id FROM vehicles WHERE id = %s;", (vehicle_id,))
        vehicle = cur.fetchone()
        if not vehicle:
            conn.close()
            return {"status": "error", "message": f"Vehículo {vehicle_id} no existe"}, 404

        # Regla de negocio: cada vehículo puede tener una única venta.
        cur.execute("SELECT id FROM sales WHERE vehicle_id = %s LIMIT 1;", (vehicle_id,))
        existing_sale = cur.fetchone()
        if existing_sale:
            conn.close()
            return {
                "status": "error",
                "message": f"El vehículo {vehicle_id} ya tiene una venta registrada"
            }, 409

        insert_columns = ["vehicle_id", price_column]
        values = [vehicle_id, amount]
        optional_fields = {
            "moneda": moneda,
            "tasa_cambio": tasa_cambio,
            date_column: fecha_venta,
            "nombre_cliente": nombre_cliente,
            "telefono_cliente": telefono_cliente,
            "metodo_pago": metodo_pago,
            "notas": notas,
        }

        for col, val in optional_fields.items():
            if col and col in sales_columns and val is not None:
                insert_columns.append(col)
                values.append(val)

        if "created_at" in sales_columns:
            insert_columns.append("created_at")
            placeholders = ["%s"] * len(values) + ["NOW()"]
        else:
            placeholders = ["%s"] * len(values)

        # Query dinámica para soportar variaciones de columnas históricas en la tabla `sales`.
        query = f"""
            INSERT INTO sales ({", ".join(insert_columns)})
            VALUES ({", ".join(placeholders)})
            RETURNING id;
        """

        cur.execute(query, tuple(values))
        new_id = cur.fetchone()[0]
        log_audit(conn, "create", "sale", new_id, {"payload": data})
        
        # Regla de negocio: cuando se registra una venta,
        # el vehículo debe salir del inventario disponible.
        cur.execute("""
            UPDATE vehicles
            SET estado = 'vendido',
                fecha_venta = COALESCE(%s, CURRENT_DATE)
            WHERE id = %s;
        """, (fecha_venta, vehicle_id))

        conn.commit()
        cur.close()
        conn.close()

        return {"status": "success", "message": "Venta creada correctamente", "id": new_id}, 201

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/sales", methods=["GET"])
def get_sales():
    """
    Lista todas las ventas registradas.

    Parámetros:
        - No recibe parámetros de ruta ni query params.

    Comportamiento:
        - Obtiene todas las filas de `sales`.
        - Ordena por fecha de venta descendente cuando la columna de fecha existe.
        - Si no existe columna de fecha, ordena por `id` descendente.
        - Normaliza la salida para exponer un contrato consistente (`precio_venta`, `fecha_venta`, etc.).

    Respuestas HTTP:
        - 200: listado de ventas.
        - 500: error interno.
    """
    try:
        conn = get_connection()
        sales_columns = get_table_columns(conn, "sales")
        date_column = get_sales_date_column(sales_columns)
        start_date, end_date, date_error = get_date_filters()
        if date_error:
            conn.close()
            return date_error

        date_filter_clause, date_filter_params = build_date_filter_clause(date_column, start_date, end_date)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if date_column:
            # Prioriza ventas más recientes usando la fecha disponible en el esquema actual.
            cur.execute(
                f"""
                SELECT *
                FROM sales
                WHERE 1=1{date_filter_clause}
                ORDER BY COALESCE({date_column}, NOW()) DESC, id DESC;
                """,
                tuple(date_filter_params)
            )
        else:
            # Fallback cuando no hay columna de fecha en la tabla.
            cur.execute("SELECT * FROM sales ORDER BY id DESC;")
        rows = cur.fetchall()

        cur.close()
        conn.close()

        data = [normalize_sale(row, sales_columns) for row in rows]
        return {"status": "success", "data": data}

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/vehicles/<int:vehicle_id>/sales", methods=["GET"])
def get_sales_by_vehicle(vehicle_id):
    """
    Lista ventas asociadas a un vehículo específico.

    Parámetros de ruta:
        - vehicle_id (int, requerido): ID del vehículo.

    Validaciones:
        - El vehículo debe existir antes de consultar sus ventas.

    Comportamiento:
        - Consulta ventas filtradas por `vehicle_id`.
        - Aplica orden descendente por fecha cuando la columna de fecha existe.
        - Si no existe columna de fecha, ordena por `id` descendente.
        - Retorna datos normalizados para mantener consistencia del contrato.

    Respuestas HTTP:
        - 200: ventas del vehículo (lista, puede estar vacía).
        - 404: vehículo no existe.
        - 500: error interno.
    """
    try:
        conn = get_connection()
        sales_columns = get_table_columns(conn, "sales")
        date_column = get_sales_date_column(sales_columns)
        start_date, end_date, date_error = get_date_filters()
        if date_error:
            conn.close()
            return date_error

        date_filter_clause, date_filter_params = build_date_filter_clause(date_column, start_date, end_date)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        # Validación explícita de existencia del recurso padre (vehículo).
        cur.execute("SELECT id FROM vehicles WHERE id = %s;", (vehicle_id,))
        vehicle = cur.fetchone()
        if not vehicle:
            conn.close()
            return {"status": "error", "message": f"Vehículo {vehicle_id} no existe"}, 404

        if date_column:
            cur.execute(f"""
                SELECT *
                FROM sales
                WHERE vehicle_id = %s{date_filter_clause}
                ORDER BY COALESCE({date_column}, NOW()) DESC, id DESC;
            """, (vehicle_id, *date_filter_params))
        else:
            cur.execute("""
                SELECT *
                FROM sales
                WHERE vehicle_id = %s
                ORDER BY id DESC;
            """, (vehicle_id,))
        rows = cur.fetchall()

        cur.close()
        conn.close()

        data = [normalize_sale(row, sales_columns) for row in rows]
        return {"status": "success", "data": data}

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/vehicles/<int:vehicle_id>/profit", methods=["GET"])
def get_profit_by_vehicle(vehicle_id):
    """
    Obtiene la ganancia real de un vehículo específico.

    Parámetros de ruta:
        - vehicle_id (int, requerido): ID del vehículo.

    Fórmulas:
        - total_venta = precio_venta / COALESCE(tasa_cambio, 1)
        - total_costos = SUM(monto / COALESCE(tasa_cambio, 1))
        - ganancia_real = total_venta - total_costos
        - margen_porcentaje = (ganancia_real / total_venta) * 100 si total_venta > 0, de lo contrario 0

    Respuestas HTTP:
        - 200: resumen de ganancia del vehículo.
        - 404: vehículo no existe.
        - 500: error interno.
    """
    try:
        conn = get_connection()
        sales_columns = get_table_columns(conn, "sales")
        sales_date_column = get_sales_date_column(sales_columns)
        start_date, end_date, date_error = get_date_filters()
        if date_error:
            conn.close()
            return date_error

        sales_filter_clause, sales_filter_params = build_date_filter_clause(sales_date_column, start_date, end_date)
        has_date_filter = bool(start_date or end_date)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        cur.execute(f"""
            SELECT
                v.id AS vehicle_id,
                v.vin,
                v.marca,
                v.modelo,
                v.anio,
                v.estado,
                v.precio_estimado,
                COALESCE(costs_agg.total_costos, 0) AS total_costos,
                COALESCE(sales_agg.total_venta, 0) AS total_venta,
                COALESCE(sales_agg.total_venta, 0) - COALESCE(costs_agg.total_costos, 0) AS ganancia_real
            FROM vehicles v
            LEFT JOIN (
                SELECT
                    vehicle_id,
                    SUM(monto / COALESCE(tasa_cambio, 1)) AS total_costos
                FROM costs
                GROUP BY vehicle_id
            ) costs_agg ON costs_agg.vehicle_id = v.id
            LEFT JOIN (
                SELECT
                    vehicle_id,
                    MAX(precio_venta / COALESCE(tasa_cambio, 1)) AS total_venta
                FROM sales
                WHERE 1=1{sales_filter_clause}
                GROUP BY vehicle_id
            ) sales_agg ON sales_agg.vehicle_id = v.id
            WHERE v.id = %s
              AND (%s = false OR sales_agg.vehicle_id IS NOT NULL);
        """, (*sales_filter_params, vehicle_id, has_date_filter))

        row = cur.fetchone()
        cur.close()
        conn.close()

        if row is None:
            return {"status": "error", "message": f"Vehículo {vehicle_id} no existe"}, 404

        return {"status": "success", "data": serialize_profit_row(row)}
    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/vehicles/profit-report", methods=["GET"])
def get_profit_report():
    """
    Lista la ganancia real de todos los vehículos.

    Campos de salida por vehículo:
        - vehicle_id, vin, marca, modelo, anio, estado
        - total_costos, total_venta, ganancia_real, margen_porcentaje

    Reglas:
        - Si no hay venta, total_venta retorna 0.
        - Si no hay costos, total_costos retorna 0.
        - Evita división por cero en margen.

    Respuestas HTTP:
        - 200: reporte completo.
        - 500: error interno.
    """
    try:
        conn = get_connection()
        sales_columns = get_table_columns(conn, "sales")
        sales_date_column = get_sales_date_column(sales_columns)
        start_date, end_date, date_error = get_date_filters()
        if date_error:
            conn.close()
            return date_error

        sales_filter_clause, sales_filter_params = build_date_filter_clause(sales_date_column, start_date, end_date)
        has_date_filter = bool(start_date or end_date)

        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        cur.execute(f"""
            SELECT
                v.id AS vehicle_id,
                v.vin,
                v.marca,
                v.modelo,
                v.anio,
                v.estado,
                v.precio_estimado,
                COALESCE(costs_agg.total_costos, 0) AS total_costos,
                COALESCE(sales_agg.total_venta, 0) AS total_venta,
                COALESCE(sales_agg.total_venta, 0) - COALESCE(costs_agg.total_costos, 0) AS ganancia_real
            FROM vehicles v
            LEFT JOIN (
                SELECT
                    vehicle_id,
                    SUM(monto / COALESCE(tasa_cambio, 1)) AS total_costos
                FROM costs
                GROUP BY vehicle_id
            ) costs_agg ON costs_agg.vehicle_id = v.id
            LEFT JOIN (
                SELECT
                    vehicle_id,
                    MAX(precio_venta / COALESCE(tasa_cambio, 1)) AS total_venta
                FROM sales
                WHERE 1=1{sales_filter_clause}
                GROUP BY vehicle_id
            ) sales_agg ON sales_agg.vehicle_id = v.id
            WHERE %s = false OR sales_agg.vehicle_id IS NOT NULL
            ORDER BY v.id DESC;
        """, (*sales_filter_params, has_date_filter))

        rows = cur.fetchall()
        cur.close()
        conn.close()

        data = [serialize_profit_row(row) for row in rows]
        return {"status": "success", "data": data}
    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/sales/<int:id>", methods=["PATCH"])
@require_authenticated_active_user
def patch_sale(id):
    """
    Actualiza parcialmente una venta existente.

    Parámetros de ruta:
        - id (int, requerido): ID de la venta a actualizar.

    Body esperado (JSON, parcial):
        - vehicle_id (int, opcional)
        - precio_venta (number, opcional)
        - monto (number, opcional; alias de precio_venta)
        - moneda (str, opcional)
        - tasa_cambio (number, opcional)
        - fecha_venta (date/datetime string, opcional)
        - fecha (date/datetime string, opcional; alias de fecha_venta)
        - nombre_cliente (str, opcional)
        - telefono_cliente (str, opcional)
        - metodo_pago (str, opcional)
        - notas (str, opcional)

    Validaciones:
        - Debe enviarse al menos un campo actualizable.
        - Si se cambia `vehicle_id`, el vehículo destino debe existir.
        - Regla de negocio: un vehículo solo puede tener una venta; evita duplicados al reasignar.

    Respuestas HTTP:
        - 200: venta actualizada correctamente.
        - 400: no se enviaron campos válidos para actualizar.
        - 404: venta no encontrada o vehículo destino inexistente.
        - 409: conflicto por duplicidad de venta para el vehículo.
        - 500: error interno.
    """
    try:
        data = request.get_json(silent=True) or {}

        conn = get_connection()
        sales_columns = get_table_columns(conn, "sales")
        price_column = get_sales_price_column(sales_columns)
        date_column = get_sales_date_column(sales_columns)

        if "fecha_venta" in data:
            _, date_error = validate_transaction_date(data.get("fecha_venta"), "fecha_venta")
            if date_error:
                conn.close()
                return date_error
        elif "fecha" in data:
            _, date_error = validate_transaction_date(data.get("fecha"), "fecha")
            if date_error:
                conn.close()
                return date_error

        fields = []
        values = []

        allowed = ["vehicle_id", "moneda", "tasa_cambio", "nombre_cliente", "telefono_cliente", "metodo_pago", "notas"]
        for f in allowed:
            if f in data and f in sales_columns:
                fields.append(f"{f} = %s")
                values.append(data[f])

        if "precio_venta" in data and price_column:
            fields.append(f"{price_column} = %s")
            values.append(data["precio_venta"])
        elif "monto" in data and price_column:
            fields.append(f"{price_column} = %s")
            values.append(data["monto"])

        if "fecha_venta" in data and date_column:
            fields.append(f"{date_column} = %s")
            values.append(data["fecha_venta"])
        elif "fecha" in data and date_column:
            fields.append(f"{date_column} = %s")
            values.append(data["fecha"])

        if "precio_venta" in data and "precio_venta" in sales_columns and not price_column:
            fields.append("precio_venta = %s")
            values.append(data["precio_venta"])

        # Se requiere al menos un campo válido para construir un UPDATE parcial.
        if not fields:
            conn.close()
            return {"status": "error", "message": "No hay datos para actualizar"}, 400

        cur = conn.cursor()

        if "vehicle_id" in data:
            # Validación referencial al cambiar el vehículo asociado a la venta.
            cur.execute("SELECT id FROM vehicles WHERE id = %s;", (data["vehicle_id"],))
            vehicle = cur.fetchone()
            if not vehicle:
                conn.close()
                return {"status": "error", "message": f"Vehículo {data['vehicle_id']} no existe"}, 404

            # Regla de negocio: no permitir dos ventas para un mismo vehículo.
            cur.execute("SELECT id FROM sales WHERE vehicle_id = %s AND id <> %s LIMIT 1;", (data["vehicle_id"], id))
            duplicate = cur.fetchone()
            if duplicate:
                conn.close()
                return {
                    "status": "error",
                    "message": f"El vehículo {data['vehicle_id']} ya tiene una venta registrada"
                }, 409

        values.append(id)
        # Query dinámica basada en los campos provistos por el cliente.
        query = f"""
            UPDATE sales
            SET {", ".join(fields)}
            WHERE id = %s
            RETURNING id;
        """
        cur.execute(query, tuple(values))
        updated = cur.fetchone()
        if updated is not None:
            log_audit(conn, "update", "sale", id, {"payload": data})

        conn.commit()
        cur.close()
        conn.close()

        if updated is None:
            return {"status": "error", "message": "Venta no encontrada"}, 404

        return {"status": "success", "message": "Venta actualizada correctamente", "id": id}

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500


@app.route("/sales/<int:id>", methods=["DELETE"])
@require_authenticated_active_user
def delete_sale(id):
    """
    Elimina una venta por su identificador.

    Parámetros de ruta:
        - id (int, requerido): ID de la venta.

    Comportamiento:
        - Elimina la venta.
        - Si la venta existe, revierte el vehículo asociado a estado disponible.
        - Limpia fecha_venta del vehículo.

    Respuestas HTTP:
        - 200: venta eliminada.
        - 404: venta no encontrada.
        - 500: error interno.
    """
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Eliminamos la venta y recuperamos el vehículo asociado.
        cur.execute("""
            DELETE FROM sales
            WHERE id = %s
            RETURNING id, vehicle_id;
        """, (id,))

        deleted = cur.fetchone()

        if deleted is None:
            conn.commit()
            cur.close()
            conn.close()
            return {"status": "error", "message": "Venta no encontrada"}, 404

        vehicle_id = deleted[1]
        log_audit(conn, "delete", "sale", id, {"vehicle_id": vehicle_id})

        # Al eliminar la venta, el vehículo vuelve a estar disponible.
        cur.execute("""
            UPDATE vehicles
            SET estado = 'disponible',
                fecha_venta = NULL
            WHERE id = %s;
        """, (vehicle_id,))

        conn.commit()
        cur.close()
        conn.close()

        return {"status": "success", "message": f"Venta {id} eliminada"}

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500

@app.route("/costs", methods=["POST"])
@require_authenticated_active_user
def create_cost():
    try:
        data = request.get_json(silent=True) or {}

        validation_error = validar_tipo_costo(data)
        if validation_error:
            return validation_error

        vehicle_id = data.get("vehicle_id")
        tipo = data.get("tipo")
        monto = data.get("monto")
        moneda = data.get("moneda", "DOP")
        tasa_cambio = data.get("tasa_cambio")
        fecha = data.get("fecha")
        descripcion = data.get("descripcion")

        _, date_error = validate_transaction_date(fecha, "fecha")
        if date_error:
            return date_error

        if vehicle_id is None or not tipo or monto is None:
            return {
                "status": "error",
                "message": "vehicle_id, tipo y monto son obligatorios"
            }, 400

        conn = get_connection()
        cur = conn.cursor()

        query = """
        INSERT INTO costs (vehicle_id, tipo, monto, moneda, tasa_cambio, fecha, descripcion)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        RETURNING id;
        """

        cur.execute(query, (
            vehicle_id,
            tipo,
            monto,
            moneda,
            tasa_cambio,
            fecha,
            descripcion
        ))

        new_id = cur.fetchone()[0]
        log_audit(conn, "create", "cost", new_id, {"payload": data})

        conn.commit()
        cur.close()
        conn.close()

        return {
            "status": "OK",
            "id": new_id
        }

    except Exception as e:
        return {"error": str(e)}, 500    


@app.route("/vehicles/<int:vehicle_id>/costs", methods=["GET"])
def get_costs_by_vehicle(vehicle_id):
    try:
        start_date, end_date, date_error = get_date_filters()
        if date_error:
            return date_error

        date_filter_clause, date_filter_params = build_date_filter_clause("fecha", start_date, end_date)

        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT
                id,
                vehicle_id,
                tipo,
                monto,
                moneda,
                tasa_cambio,
                fecha,
                descripcion
            FROM costs
            WHERE vehicle_id = %s
        """ + date_filter_clause + """
            ORDER BY COALESCE(fecha, NOW()) DESC, id DESC;
        """, (vehicle_id, *date_filter_params))

        rows = cur.fetchall()

        cur.close()
        conn.close()

        data = []
        for row in rows:
            data.append({
                "id": row[0],
                "vehicle_id": row[1],
                "tipo": row[2],
                "monto": float(row[3]) if row[3] is not None else 0,
                "moneda": row[4],
                "tasa_cambio": float(row[5]) if row[5] is not None else None,
                "fecha": row[6],
                "descripcion": row[7]
            })

        return {"status": "OK", "data": data}

    except Exception as e:
        return {"error": str(e)}, 500

@app.route("/costs/<int:id>", methods=["DELETE"])
@require_authenticated_active_user
def delete_cost(id):
    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("DELETE FROM costs WHERE id = %s RETURNING id;", (id,))
        deleted = cur.fetchone()
        if deleted is not None:
            log_audit(conn, "delete", "cost", id, {"deleted_id": id})

        conn.commit()
        cur.close()
        conn.close()

        if deleted is None:
            return {"error": "Costo no encontrado"}, 404

        return {"status": "OK", "message": f"Costo {id} eliminado"}

    except Exception as e:
        return {"error": str(e)}, 500
@app.route("/costs/<int:id>", methods=["PATCH"])
@require_authenticated_active_user
def patch_cost(id):
    try:
        data = request.get_json(silent=True) or {}

        print("PATCH COST ID:", id)
        print("PATCH COST DATA:", data)

        validation_error = validar_tipo_costo(data)
        if validation_error:
            print("VALIDATION ERROR:", validation_error)
            return validation_error

        fields = []
        values = []

        if "fecha" in data:
            _, date_error = validate_transaction_date(data.get("fecha"), "fecha")
            if date_error:
                return date_error

        allowed = ["tipo", "monto", "moneda", "tasa_cambio", "fecha", "descripcion"]

        for f in allowed:
            if f in data:
                fields.append(f"{f} = %s")
                values.append(data[f])

        print("FIELDS:", fields)
        print("VALUES BEFORE ID:", values)

        if not fields:
            return {
                "status": "error",
                "message": "No hay datos para actualizar"
            }, 400

        values.append(id)

        conn = get_connection()
        cur = conn.cursor()

        query = f"""
        UPDATE costs
        SET {", ".join(fields)}
        WHERE id = %s
        RETURNING id;
        """

        print("QUERY:", query)
        print("VALUES FINAL:", values)

        cur.execute(query, tuple(values))
        updated = cur.fetchone()
        if updated is not None:
            log_audit(conn, "update", "cost", id, {"payload": data})

        conn.commit()
        cur.close()
        conn.close()

        print("UPDATED:", updated)

        if updated is None:
            return {
                "status": "error",
                "message": "Costo no encontrado"
            }, 404

        return {
            "status": "OK",
            "message": f"Costo {id} actualizado"
        }

    except Exception as e:
        print("PATCH COST ERROR:", str(e))
        return {"error": str(e)}, 500

@app.route("/vehicles/<int:vehicle_id>/costs/total", methods=["GET"])
def total_costs(vehicle_id):
    try:
        start_date, end_date, date_error = get_date_filters()
        if date_error:
            return date_error

        date_filter_clause, date_filter_params = build_date_filter_clause("fecha", start_date, end_date)

        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT COALESCE(SUM(monto / COALESCE(tasa_cambio, 1)), 0)
            FROM costs
            WHERE vehicle_id = %s
        """ + date_filter_clause, (vehicle_id, *date_filter_params))

        total = cur.fetchone()[0]

        cur.close()
        conn.close()

        return {"total_cost": float(total)}

    except Exception as e:
        return {"error": str(e)}, 500            

@app.route("/vehicles/profit-real", methods=["GET"])
def vehicles_real_profit():
    try:
        conn = get_connection()
        sales_columns = get_table_columns(conn, "sales")
        sales_date_column = get_sales_date_column(sales_columns)
        start_date, end_date, date_error = get_date_filters()
        if date_error:
            conn.close()
            return date_error

        sales_filter_clause, sales_filter_params = build_date_filter_clause(sales_date_column, start_date, end_date, "s")
        has_date_filter = bool(start_date or end_date)

        cur = conn.cursor()

        query = f"""
        WITH costs_by_vehicle AS (
            SELECT
                vehicle_id,
                COALESCE(SUM(monto / COALESCE(tasa_cambio, 1)), 0) AS total_costos
            FROM costs
            GROUP BY vehicle_id
        ),
        sales_by_vehicle AS (
            SELECT
                s.vehicle_id,
                COALESCE(SUM(
                    COALESCE(
                        NULLIF(to_jsonb(s) ->> 'monto', '')::numeric,
                        NULLIF(to_jsonb(s) ->> 'precio_venta', '')::numeric,
                        0
                    ) * COALESCE(
                        NULLIF(to_jsonb(s) ->> 'tasa_cambio', '')::numeric,
                        1
                    )
                ), 0) AS total_ventas
            FROM sales s
            WHERE 1=1{sales_filter_clause}
            GROUP BY s.vehicle_id
        )
        SELECT
            v.id,
            v.vin,
            v.marca,
            v.modelo,
            v.anio,
            COALESCE(sbv.total_ventas, 0) AS total_ventas,
            COALESCE(cbv.total_costos, 0) AS total_costos,
            COALESCE(sbv.total_ventas, 0) - COALESCE(cbv.total_costos, 0) AS ganancia_real
        FROM vehicles v
        LEFT JOIN costs_by_vehicle cbv ON cbv.vehicle_id = v.id
        LEFT JOIN sales_by_vehicle sbv ON sbv.vehicle_id = v.id
        WHERE %s = false OR sbv.vehicle_id IS NOT NULL
        ORDER BY v.id;
        """

        cur.execute(query, (*sales_filter_params, has_date_filter))
        result = cur.fetchall()

        cur.close()
        conn.close()

        data = []
        for row in result:
            data.append({
                "vehicle_id": row[0],
                "vin": row[1],
                "marca": row[2],
                "modelo": row[3],
                "anio": row[4],
                "total_ventas": float(row[5]),
                "total_costos": float(row[6]),
                "ganancia_real": float(row[7]),
            })

        return {"status": "OK", "data": data}

    except Exception as e:
        return {"error": str(e)}, 500

@app.route("/debug/db", methods=["GET"])
def debug_db():
    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("SELECT current_database(), inet_server_port();")
        db_info = cur.fetchone()

        cur.execute("""
            SELECT id, vehicle_id, precio_venta, fecha_venta, nombre_cliente
            FROM sales
            ORDER BY id;
        """)
        sales = cur.fetchall()

        cur.close()
        conn.close()

        return {
            "database": db_info[0],
            "port": db_info[1],
            "sales": [
                {
                    "id": r[0],
                    "vehicle_id": r[1],
                    "precio_venta": float(r[2]) if r[2] is not None else None,
                    "fecha_venta": str(r[3]) if r[3] is not None else None,
                    "nombre_cliente": r[4]
                }
                for r in sales
            ]
        }

    except Exception as e:
        return {"error": str(e)}, 500        


if __name__ == "__main__":
    Swagger(app)
    app.run(debug=app.config["FLASK_ENV"] != "production")
